import re
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import load_workbook

EMAIL_REGEX = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Header names (case-insensitive) we recognize for optional contact fields.
_NAME_HEADERS = {"name", "full name", "contact name"}
_UNIVERSITY_HEADERS = {"university", "institution", "school"}


def extract_contacts_from_excel(file_bytes: bytes) -> List[Dict[str, Optional[str]]]:
    """Read an uploaded Excel file and return a deduplicated list of contacts.

    Each contact is {"email": str, "name": str | None, "university": str | None}.

    Looks for a column named "email" (case-insensitive) in the header row of
    the first sheet, plus optional "name" and "university" columns. If no
    "email" column exists, falls back to scanning every cell in the sheet
    and treating anything that looks like a valid email address as a contact
    with no name/university.
    """
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return []

    email_col = name_col = university_col = None
    for i, cell in enumerate(header):
        if cell is None:
            continue
        label = str(cell).strip().lower()
        if label == "email":
            email_col = i
        elif label in _NAME_HEADERS:
            name_col = i
        elif label in _UNIVERSITY_HEADERS:
            university_col = i

    contacts: List[Dict[str, Optional[str]]] = []
    seen = set()

    def _cell(row, idx):
        if idx is None or idx >= len(row) or row[idx] is None:
            return None
        value = str(row[idx]).strip()
        return value or None

    if email_col is not None:
        for row in rows:
            email = _cell(row, email_col)
            if email and EMAIL_REGEX.match(email) and email.lower() not in seen:
                seen.add(email.lower())
                contacts.append(
                    {"email": email, "name": _cell(row, name_col), "university": _cell(row, university_col)}
                )
    else:
        # No "email" header — scan every cell in the sheet, including the header row.
        for row in sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                candidate = str(cell).strip()
                if EMAIL_REGEX.match(candidate) and candidate.lower() not in seen:
                    seen.add(candidate.lower())
                    contacts.append({"email": candidate, "name": None, "university": None})

    return contacts
